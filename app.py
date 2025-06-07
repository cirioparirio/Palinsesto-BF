import streamlit as st
import pandas as pd
import numpy as np
import base64
from io import BytesIO
import os
import pickle
import tempfile
import re
import datetime

# Configurazione della pagina Streamlit
st.set_page_config(
    page_title="Visualizzatore Palinsesto BF",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Funzione di supporto per confrontare gli orari
def self_compare_time(time_str, min_hour, min_minute, max_hour, max_minute):
    """
    Confronta un orario in formato stringa con un intervallo di orari
    """
    try:
        # Estrai ore e minuti dalla stringa
        time_match = re.search(r'(\d{1,2}):(\d{2})', str(time_str))
        if time_match:
            hour = int(time_match.group(1))
            minute = int(time_match.group(2))
            
            # Converti in minuti totali per un confronto più semplice
            time_minutes = hour * 60 + minute
            min_minutes = min_hour * 60 + min_minute
            max_minutes = max_hour * 60 + max_minute
            
            # Confronta con l'intervallo
            return min_minutes <= time_minutes <= max_minutes
    except Exception as e:
        # Se c'è un errore, stampa il messaggio di errore ma mantieni la riga
        return True
    
    # Se il formato non corrisponde, mantieni la riga
    return True

# Funzione per salvare i dati in memoria
def save_data(df, color_df, file_name, file_type="palinsesto"):
    """
    Salva i dati in un file pickle per la persistenza
    """
    # Crea una directory temporanea se non esiste
    if not os.path.exists(".streamlit"):
        os.makedirs(".streamlit")
    
    # Salva il DataFrame e il nome del file
    with open(f".streamlit/data_{file_type}.pkl", "wb") as f:
        pickle.dump({"df": df, "color_df": color_df, "file_name": file_name}, f)

# Funzione per caricare i dati dalla memoria
def load_data(file_type="palinsesto"):
    """
    Carica i dati dal file pickle se esiste
    """
    if os.path.exists(f".streamlit/data_{file_type}.pkl"):
        try:
            with open(f".streamlit/data_{file_type}.pkl", "rb") as f:
                data = pickle.load(f)
            return data["df"], data["color_df"], data["file_name"]
        except Exception as e:
            st.error(f"Errore durante il caricamento dei dati salvati: {e}")
    return None, None, None

# Funzione per generare un link di download per un DataFrame
def get_table_download_link(df, filename, text):
    """Genera un link per scaricare il dataframe come file CSV"""
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="{filename}">📥 {text}</a>'
    return href

# Funzione per caricare il file Excel
def load_excel_file(uploaded_file, sheet_name=None):
    try:
        # Crea un file temporaneo per salvare il file caricato
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        
        # Se è specificato un nome di foglio, carica quel foglio specifico
        if sheet_name:
            # Leggi il file Excel con openpyxl per preservare i colori delle celle
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            
            # Verifica se il foglio esiste
            if sheet_name not in wb.sheetnames:
                st.error(f"Il foglio '{sheet_name}' non esiste nel file Excel")
                os.unlink(tmp_path)
                return None, None
            
            sheet = wb[sheet_name]
            
            # Estrai i dati e le informazioni sui colori
            data = []
            colors = []
            
            for row in sheet.iter_rows():
                row_data = []
                row_colors = []
                for cell in row:
                    row_data.append(cell.value)
                    # Estrai il colore di sfondo della cella
                    fill_color = cell.fill.start_color.index
                    row_colors.append(fill_color)
                data.append(row_data)
                colors.append(row_colors)
            
            # Crea un DataFrame dai dati
            if data:
                # Usa la prima riga come intestazioni
                headers = data[0]
                
                # Verifica se ci sono colonne duplicate e rinominale
                header_counts = {}
                unique_headers = []
                for h in headers:
                    if h in header_counts:
                        header_counts[h] += 1
                        unique_headers.append(f"{h}_{header_counts[h]}")
                    else:
                        header_counts[h] = 0
                        unique_headers.append(h)
                
                # Crea il DataFrame con le intestazioni uniche
                df = pd.DataFrame(data[1:], columns=unique_headers)
                
                # Salva le informazioni sui colori nel DataFrame
                color_df = pd.DataFrame(colors[1:], columns=unique_headers)
                
                # Pulisci il file temporaneo
                os.unlink(tmp_path)
                
                return df, color_df
            else:
                st.error("Il foglio Excel è vuoto")
                os.unlink(tmp_path)
                return None, None
        else:
            # Carica il foglio predefinito
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            sheet = wb.active
            
            # Estrai i dati e le informazioni sui colori
            data = []
            colors = []
            
            for row in sheet.iter_rows():
                row_data = []
                row_colors = []
                for cell in row:
                    row_data.append(cell.value)
                    # Estrai il colore di sfondo della cella
                    fill_color = cell.fill.start_color.index
                    row_colors.append(fill_color)
                data.append(row_data)
                colors.append(row_colors)
            
            # Crea un DataFrame dai dati
            if data:
                # Usa la prima riga come intestazioni
                headers = data[0]
                
                # Verifica se ci sono colonne duplicate e rinominale
                header_counts = {}
                unique_headers = []
                for h in headers:
                    if h in header_counts:
                        header_counts[h] += 1
                        unique_headers.append(f"{h}_{header_counts[h]}")
                    else:
                        header_counts[h] = 0
                        unique_headers.append(h)
                
                # Crea il DataFrame con le intestazioni uniche
                df = pd.DataFrame(data[1:], columns=unique_headers)
                
                # Salva le informazioni sui colori nel DataFrame
                color_df = pd.DataFrame(colors[1:], columns=unique_headers)
                
                # Pulisci il file temporaneo
                os.unlink(tmp_path)
                
                return df, color_df
            else:
                st.error("Il file Excel è vuoto")
                os.unlink(tmp_path)
                return None, None
            
    except Exception as e:
        st.error(f"Errore durante il caricamento del file: {e}")
        return None, None

# Funzione per formattare la colonna ORA in formato hh:mm
def format_time_column(df, column_name='ORA'):
    """
    Formatta la colonna ORA in formato hh:mm
    """
    if column_name in df.columns:
        # Crea una copia per evitare SettingWithCopyWarning
        df_copy = df.copy()
        
        # Applica la formattazione a tutte le celle della colonna ORA
        for i in df_copy.index:
            value = df_copy.at[i, column_name]
            if pd.notna(value):  # Verifica che il valore non sia NaN
                # Gestisci diversi tipi di dati per l'ora
                if isinstance(value, datetime.time):
                    # Se è un oggetto datetime.time, formatta direttamente
                    df_copy.at[i, column_name] = f"{value.hour:02d}:{value.minute:02d}"
                else:
                    # Converti il valore in stringa
                    value_str = str(value)
                    
                    # Cerca un pattern di orario nel formato hh:mm:ss
                    time_match = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', value_str)
                    if time_match:
                        # Estrai ore e minuti
                        hours, minutes = time_match.group(1), time_match.group(2)
                        # Formatta come hh:mm
                        df_copy.at[i, column_name] = f"{hours}:{minutes}"
        
        return df_copy
    return df

# Funzione per formattare i numeri con virgole e un decimale
def format_numeric_columns(df, start_col=5, end_col=None):
    """
    Formatta le colonne numeriche con virgole e un decimale
    """
    # Determina l'intervallo di colonne da formattare
    if end_col is None:
        # Se end_col non è specificato, usa tutte le colonne a partire da start_col
        numeric_cols = df.columns[start_col:]
    else:
        # Altrimenti, usa l'intervallo specificato
        numeric_cols = df.columns[start_col:end_col]
    
    # Crea una copia per evitare SettingWithCopyWarning
    df_copy = df.copy()
    
    # Formatta le colonne numeriche
    for col in numeric_cols:
        if col in df.columns:
            # Converti i valori in stringhe con un decimale e virgola
            # Usa un blocco try-except per gestire eventuali errori
            try:
                df_copy[col] = df_copy[col].apply(
                    lambda x: str(x).replace('.', ',') if pd.notnull(x) and isinstance(x, (int, float)) else x
                )
            except Exception as e:
                # Se c'è un errore, mantieni la colonna originale
                st.warning(f"Impossibile formattare la colonna {col}: {e}")
                continue
    
    return df_copy

# Funzione per applicare lo stile alle celle colorate
def apply_cell_styling(df, color_df):
    """
    Applica lo stile alle celle del DataFrame per visualizzare i colori
    Ritorna un DataFrame con stile applicato
    """
    # Verifica che i DataFrame abbiano le stesse dimensioni
    if df.shape[0] != color_df.shape[0]:
        # Se hanno dimensioni diverse, ridimensiona color_df per adattarlo a df
        if df.shape[0] > color_df.shape[0]:
            # Aggiungi righe vuote a color_df
            rows_to_add = df.shape[0] - color_df.shape[0]
            empty_rows = pd.DataFrame([[0] * len(color_df.columns)] * rows_to_add, columns=color_df.columns)
            color_df = pd.concat([color_df, empty_rows], ignore_index=True)
        else:
            # Taglia color_df per adattarlo a df
            color_df = color_df.iloc[:df.shape[0]]
    
    # Verifica che le colonne corrispondano
    common_cols = [col for col in df.columns if col in color_df.columns]
    
    # Applica lo stile
    styled_df = pd.DataFrame('', index=df.index, columns=df.columns)
    
    for col in common_cols:
        for i in range(len(df)):
            try:
                if i < len(color_df) and color_df[col].iloc[i] != '00000000' and color_df[col].iloc[i] != 0:
                    styled_df.loc[i, col] = 'background-color: #90EE90'
            except Exception as e:
                # Se c'è un errore, ignora questa cella
                continue
    
    return df.style.apply(lambda _: styled_df, axis=None)

# Inizializza lo stato della sessione
if 'palinsesto_loaded' not in st.session_state:
    st.session_state.palinsesto_loaded = False
    st.session_state.palinsesto_df = None
    st.session_state.palinsesto_color_df = None
    st.session_state.palinsesto_file_name = None

if 'archivio_loaded' not in st.session_state:
    st.session_state.archivio_loaded = False
    st.session_state.archivio_df = None
    st.session_state.archivio_color_df = None
    st.session_state.archivio_file_name = None

# Titolo dell'applicazione
st.title("📊 Visualizzatore Palinsesto BF")
st.markdown("""
Questa applicazione permette di visualizzare e filtrare i dati.
I dati caricati rimangono in memoria anche dopo la chiusura del browser.
""")

# Sidebar per le opzioni
st.sidebar.header("Opzioni")

# Selezione del tipo di dati da visualizzare
data_type = st.sidebar.radio(
    "Seleziona il tipo di dati da visualizzare:",
    ["Palinsesto BF", "Giornata Odierna FB"]
)

# Caricamento del file Excel per il palinsesto
if data_type == "Palinsesto BF":
    uploaded_file = st.sidebar.file_uploader("Carica un nuovo file Excel per il Palinsesto", type=["xlsx"], key="palinsesto_uploader")
    
    # Carica i dati esistenti o usa quelli appena caricati
    if uploaded_file is not None:
        # Se viene caricato un nuovo file, sostituisci i dati esistenti
        df, color_df = load_excel_file(uploaded_file)
        if df is not None:
            # Applica le formattazioni richieste
            df = format_time_column(df)
            df = format_numeric_columns(df, start_col=5, end_col=17)
            
            st.session_state.palinsesto_df = df
            st.session_state.palinsesto_color_df = color_df
            st.session_state.palinsesto_file_name = uploaded_file.name
            st.session_state.palinsesto_loaded = True
            
            # Salva i dati per la persistenza
            save_data(df, color_df, uploaded_file.name, "palinsesto")
            
            st.sidebar.success(f"File caricato con successo")
    elif not st.session_state.palinsesto_loaded:
        # Carica i dati salvati se non ci sono dati nella sessione
        df, color_df, file_name = load_data("palinsesto")
        if df is not None:
            # Applica le formattazioni richieste
            df = format_time_column(df)
            df = format_numeric_columns(df, start_col=5, end_col=17)
            
            st.session_state.palinsesto_df = df
            st.session_state.palinsesto_color_df = color_df
            st.session_state.palinsesto_file_name = file_name
            st.session_state.palinsesto_loaded = True
            st.sidebar.info(f"Dati caricati dal file salvato")
        else:
            st.info("Nessun dato caricato. Carica un file Excel per iniziare.")
    
    # Se ci sono dati da visualizzare
    if st.session_state.palinsesto_loaded and st.session_state.palinsesto_df is not None:
        df = st.session_state.palinsesto_df
        color_df = st.session_state.palinsesto_color_df
        
        # Mostra informazioni sul DataFrame
        st.sidebar.write(f"Numero di righe: {df.shape[0]}")
        st.sidebar.write(f"Numero di colonne: {df.shape[1]}")
        
        # Opzioni di filtro
        st.sidebar.header("Filtri")
        
        # Crea filtri per ogni colonna
        filtered_df = df.copy()
        
        # Raggruppa le colonne in categorie per una migliore organizzazione
        # Mostra solo alcune colonne nel sidebar per non sovraffollarlo
        if len(df.columns) > 10:
            # Seleziona le colonne da mostrare nel sidebar
            filter_columns = st.sidebar.multiselect(
                "Seleziona le colonne da filtrare",
                options=list(df.columns),
                default=list(df.columns)[:5]  # Default: prime 5 colonne
            )
        else:
            filter_columns = list(df.columns)
        
        # Crea filtri per le colonne selezionate
        for column in filter_columns:
            # Ottieni i valori unici nella colonna
            unique_values = df[column].dropna().unique()
            
            # Se ci sono troppi valori unici, usa un input di testo invece di un multiselect
            if len(unique_values) > 10:
                filter_type = st.sidebar.radio(
                    f"Tipo di filtro per {column}",
                    ["Testo", "Range"],
                    key=f"filter_type_{column}"
                )
                
                if filter_type == "Testo":
                    filter_value = st.sidebar.text_input(f"Filtra {column} (testo)", key=f"text_{column}")
                    if filter_value:
                        filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(filter_value, case=False)]
                else:  # Range
                    # Gestione speciale per la colonna ORA
                    if column == 'ORA':
                        # Crea un input di testo per l'ora minima e massima
                        min_time = st.sidebar.text_input(f"Ora minima (formato hh:mm)", value="00:00", key=f"min_time_{column}")
                        max_time = st.sidebar.text_input(f"Ora massima (formato hh:mm)", value="23:59", key=f"max_time_{column}")
                        
                        # Filtra in base all'ora
                        if min_time and max_time:
                            # Converti le stringhe in oggetti time per il confronto
                            try:
                                min_time_parts = min_time.split(':')
                                max_time_parts = max_time.split(':')
                                
                                if len(min_time_parts) >= 2 and len(max_time_parts) >= 2:
                                    min_hour, min_minute = int(min_time_parts[0]), int(min_time_parts[1])
                                    max_hour, max_minute = int(max_time_parts[0]), int(max_time_parts[1])
                                    
                                    # Filtra le righe in base all'ora
                                    filtered_rows = []
                                    for i, row in filtered_df.iterrows():
                                        time_str = str(row[column])
                                        if self_compare_time(time_str, min_hour, min_minute, max_hour, max_minute):
                                            filtered_rows.append(i)
                                    
                                    filtered_df = filtered_df.loc[filtered_rows]
                            except Exception as e:
                                st.warning(f"Formato ora non valido. Usa il formato hh:mm. Errore: {e}")
                    else:
                        # Determina il tipo di dati della colonna
                        col_dtype = df[column].dtype
                        
                        if np.issubdtype(col_dtype, np.number):
                            # Per colonne numeriche
                            min_val = float(df[column].min()) if not pd.isna(df[column].min()) else 0
                            max_val = float(df[column].max()) if not pd.isna(df[column].max()) else 100
                            
                            range_min, range_max = st.sidebar.slider(
                                f"Intervallo per {column}",
                                min_value=min_val,
                                max_value=max_val,
                                value=(min_val, max_val),
                                key=f"range_{column}"
                            )
                            
                            filtered_df = filtered_df[(filtered_df[column] >= range_min) & (filtered_df[column] <= range_max)]
                        elif pd.api.types.is_datetime64_any_dtype(col_dtype):
                            # Per colonne di data
                            min_date = df[column].min().date() if not pd.isna(df[column].min()) else datetime.date.today()
                            max_date = df[column].max().date() if not pd.isna(df[column].max()) else datetime.date.today()
                            
                            date_min, date_max = st.sidebar.date_input(
                                f"Intervallo di date per {column}",
                                value=[min_date, max_date],
                                key=f"date_range_{column}"
                            )
                            
                            if isinstance(date_min, list) and len(date_min) == 2:  # Se sono state selezionate due date
                                date_min, date_max = date_min
                                filtered_df = filtered_df[(filtered_df[column].dt.date >= date_min) & (filtered_df[column].dt.date <= date_max)]
                            elif isinstance(date_min, datetime.date) and isinstance(date_max, datetime.date):
                                filtered_df = filtered_df[(filtered_df[column].dt.date >= date_min) & (filtered_df[column].dt.date <= date_max)]
                        else:
                            # Per altri tipi di colonne, usa il filtro di testo
                            filter_value = st.sidebar.text_input(f"Filtra {column} (testo)", key=f"text_fallback_{column}")
                            if filter_value:
                                filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(filter_value, case=False)]
            else:
                # Altrimenti usa un multiselect
                selected_values = st.sidebar.multiselect(
                    f"Filtra {column}",
                    options=sorted(unique_values),
                    default=[],
                    key=f"multiselect_{column}"
                )
                if selected_values:
                    filtered_df = filtered_df[filtered_df[column].isin(selected_values)]
        
        # Pulsante per reimpostare i filtri
        if st.sidebar.button("Reimposta filtri", key="reset_palinsesto"):
            filtered_df = df.copy()
        
        # Opzioni di visualizzazione
        st.header("Dati del Palinsesto")
        
        # Mostra il numero di righe filtrate
        st.write(f"Visualizzazione di {filtered_df.shape[0]} righe su {df.shape[0]} totali")
        
        # Aggiungi un link per scaricare i dati filtrati
        st.markdown(get_table_download_link(filtered_df, "palinsesto_filtrato.csv", "Scarica i dati filtrati come CSV"), unsafe_allow_html=True)
        
        # Visualizza il DataFrame filtrato con colori
        try:
            if color_df is not None:
                st.dataframe(apply_cell_styling(filtered_df, color_df), use_container_width=True)
            else:
                st.dataframe(filtered_df, use_container_width=True)
        except Exception as e:
            st.error(f"Errore durante la visualizzazione dei dati: {e}")
            # Fallback alla visualizzazione senza stile
            st.dataframe(filtered_df, use_container_width=True)

# Caricamento del file Excel per l'archivio
else:  # data_type == "Giornata Odierna FB"
    uploaded_file = st.sidebar.file_uploader("Carica un nuovo file Excel per la Giornata Odierna", type=["xlsx"], key="archivio_uploader")
    
    # Carica i dati esistenti o usa quelli appena caricati
    if uploaded_file is not None:
        # Se viene caricato un nuovo file, sostituisci i dati esistenti
        df, color_df = load_excel_file(uploaded_file, sheet_name="Giornata Odierna")
        if df is not None:
            try:
                # Applica le formattazioni richieste
                df = format_time_column(df, column_name='ORA')
                
                # Converti le colonne numeriche in modo sicuro
                # Prima converti le colonne in numeri dove possibile
                for col in df.columns[9:]:  # Colonne da J a BD
                    try:
                        # Tenta di convertire la colonna in numerica
                        df[col] = pd.to_numeric(df[col], errors='ignore')
                    except:
                        # Se fallisce, lascia la colonna com'è
                        pass
                
                # Poi applica la formattazione
                df = format_numeric_columns(df, start_col=9)  # Colonne da J a BD
                
                st.session_state.archivio_df = df
                st.session_state.archivio_color_df = color_df
                st.session_state.archivio_file_name = uploaded_file.name
                st.session_state.archivio_loaded = True
                
                # Salva i dati per la persistenza
                save_data(df, color_df, uploaded_file.name, "archivio")
                
                st.sidebar.success(f"File caricato con successo")
            except Exception as e:
                st.error(f"Errore durante la formattazione dei dati: {e}")
    elif not st.session_state.archivio_loaded:
        # Carica i dati salvati se non ci sono dati nella sessione
        df, color_df, file_name = load_data("archivio")
        if df is not None:
            try:
                # Applica le formattazioni richieste
                df = format_time_column(df, column_name='ORA')
                
                # Converti le colonne numeriche in modo sicuro
                # Prima converti le colonne in numeri dove possibile
                for col in df.columns[9:]:  # Colonne da J a BD
                    try:
                        # Tenta di convertire la colonna in numerica
                        df[col] = pd.to_numeric(df[col], errors='ignore')
                    except:
                        # Se fallisce, lascia la colonna com'è
                        pass
                
                # Poi applica la formattazione
                df = format_numeric_columns(df, start_col=9)  # Colonne da J a BD
                
                st.session_state.archivio_df = df
                st.session_state.archivio_color_df = color_df
                st.session_state.archivio_file_name = file_name
                st.session_state.archivio_loaded = True
                st.sidebar.info(f"Dati caricati dal file salvato")
            except Exception as e:
                st.error(f"Errore durante la formattazione dei dati salvati: {e}")
        else:
            st.info("Nessun dato caricato. Carica un file Excel per iniziare.")
    
    # Se ci sono dati da visualizzare
    if st.session_state.archivio_loaded and st.session_state.archivio_df is not None:
        df = st.session_state.archivio_df
        color_df = st.session_state.archivio_color_df
        
        # Mostra informazioni sul DataFrame
        st.sidebar.write(f"Numero di righe: {df.shape[0]}")
        st.sidebar.write(f"Numero di colonne: {df.shape[1]}")
        
        # Opzioni di filtro
        st.sidebar.header("Filtri")
        
        # Crea filtri per ogni colonna
        filtered_df = df.copy()
        
        # Raggruppa le colonne in categorie per una migliore organizzazione
        # Mostra solo alcune colonne nel sidebar per non sovraffollarlo
        if len(df.columns) > 10:
            # Seleziona le colonne da mostrare nel sidebar
            filter_columns = st.sidebar.multiselect(
                "Seleziona le colonne da filtrare",
                options=list(df.columns),
                default=list(df.columns)[:5]  # Default: prime 5 colonne
            )
        else:
            filter_columns = list(df.columns)
        
        # Crea filtri per le colonne selezionate
        for column in filter_columns:
            # Ottieni i valori unici nella colonna
            unique_values = df[column].dropna().unique()
            
            # Se ci sono troppi valori unici, usa un input di testo invece di un multiselect
            if len(unique_values) > 10:
                filter_type = st.sidebar.radio(
                    f"Tipo di filtro per {column}",
                    ["Testo", "Range"],
                    key=f"filter_type_arch_{column}"
                )
                
                if filter_type == "Testo":
                    filter_value = st.sidebar.text_input(f"Filtra {column} (testo)", key=f"text_arch_{column}")
                    if filter_value:
                        filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(filter_value, case=False)]
                else:  # Range
                    # Gestione speciale per la colonna ORA
                    if column == 'ORA':
                        # Crea un input di testo per l'ora minima e massima
                        min_time = st.sidebar.text_input(f"Ora minima (formato hh:mm)", value="00:00", key=f"min_time_arch_{column}")
                        max_time = st.sidebar.text_input(f"Ora massima (formato hh:mm)", value="23:59", key=f"max_time_arch_{column}")
                        
                        # Filtra in base all'ora
                        if min_time and max_time:
                            # Converti le stringhe in oggetti time per il confronto
                            try:
                                min_time_parts = min_time.split(':')
                                max_time_parts = max_time.split(':')
                                
                                if len(min_time_parts) >= 2 and len(max_time_parts) >= 2:
                                    min_hour, min_minute = int(min_time_parts[0]), int(min_time_parts[1])
                                    max_hour, max_minute = int(max_time_parts[0]), int(max_time_parts[1])
                                    
                                    # Filtra le righe in base all'ora
                                    filtered_rows = []
                                    for i, row in filtered_df.iterrows():
                                        time_str = str(row[column])
                                        if self_compare_time(time_str, min_hour, min_minute, max_hour, max_minute):
                                            filtered_rows.append(i)
                                    
                                    filtered_df = filtered_df.loc[filtered_rows]
                            except Exception as e:
                                st.warning(f"Formato ora non valido. Usa il formato hh:mm. Errore: {e}")
                    else:
                        # Determina il tipo di dati della colonna
                        col_dtype = df[column].dtype
                        
                        if np.issubdtype(col_dtype, np.number):
                            # Per colonne numeriche
                            min_val = float(df[column].min()) if not pd.isna(df[column].min()) else 0
                            max_val = float(df[column].max()) if not pd.isna(df[column].max()) else 100
                            
                            range_min, range_max = st.sidebar.slider(
                                f"Intervallo per {column}",
                                min_value=min_val,
                                max_value=max_val,
                                value=(min_val, max_val),
                                key=f"range_arch_{column}"
                            )
                            
                            filtered_df = filtered_df[(filtered_df[column] >= range_min) & (filtered_df[column] <= range_max)]
                        elif pd.api.types.is_datetime64_any_dtype(col_dtype):
                            # Per colonne di data
                            min_date = df[column].min().date() if not pd.isna(df[column].min()) else datetime.date.today()
                            max_date = df[column].max().date() if not pd.isna(df[column].max()) else datetime.date.today()
                            
                            date_min, date_max = st.sidebar.date_input(
                                f"Intervallo di date per {column}",
                                value=[min_date, max_date],
                                key=f"date_range_arch_{column}"
                            )
                            
                            if isinstance(date_min, list) and len(date_min) == 2:  # Se sono state selezionate due date
                                date_min, date_max = date_min
                                filtered_df = filtered_df[(filtered_df[column].dt.date >= date_min) & (filtered_df[column].dt.date <= date_max)]
                            elif isinstance(date_min, datetime.date) and isinstance(date_max, datetime.date):
                                filtered_df = filtered_df[(filtered_df[column].dt.date >= date_min) & (filtered_df[column].dt.date <= date_max)]
                        else:
                            # Per altri tipi di colonne, usa il filtro di testo
                            filter_value = st.sidebar.text_input(f"Filtra {column} (testo)", key=f"text_fallback_arch_{column}")
                            if filter_value:
                                filtered_df = filtered_df[filtered_df[column].astype(str).str.contains(filter_value, case=False)]
            else:
                # Altrimenti usa un multiselect
                selected_values = st.sidebar.multiselect(
                    f"Filtra {column}",
                    options=sorted(unique_values),
                    default=[],
                    key=f"multiselect_arch_{column}"
                )
                if selected_values:
                    filtered_df = filtered_df[filtered_df[column].isin(selected_values)]
        
        # Pulsante per reimpostare i filtri
        if st.sidebar.button("Reimposta filtri", key="reset_archivio"):
            filtered_df = df.copy()
        
        # Opzioni di visualizzazione
        st.header("Dati della Giornata Odierna")
        
        # Mostra il numero di righe filtrate
        st.write(f"Visualizzazione di {filtered_df.shape[0]} righe su {df.shape[0]} totali")
        
        # Aggiungi un link per scaricare i dati filtrati
        st.markdown(get_table_download_link(filtered_df, "giornata_odierna_filtrata.csv", "Scarica i dati filtrati come CSV"), unsafe_allow_html=True)
        
        # Visualizza il DataFrame filtrato con colori
        try:
            if color_df is not None:
                st.dataframe(apply_cell_styling(filtered_df, color_df), use_container_width=True)
            else:
                st.dataframe(filtered_df, use_container_width=True)
        except Exception as e:
            st.error(f"Errore durante la visualizzazione dei dati: {e}")
            # Fallback alla visualizzazione senza stile
            st.dataframe(filtered_df, use_container_width=True)

